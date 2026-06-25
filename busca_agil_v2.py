import html as html_lib
import os
import re
import tempfile
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Tuple

import gradio as gr
import openpyxl
import pdfplumber
from openpyxl.styles import Font, PatternFill

css = """
html, body, .gradio-container { margin: 0; }
.main-row { display: flex; padding: 10px; }
#controls-col { max-width: 600px; padding-right: 20px; box-sizing: border-box; }
#results-col { flex: 1; }
#table-container { height: 80vh; overflow-y: auto; }
.results-table { width: 100%; border-collapse: collapse; }
.results-table th, .results-table td {
  border: 1px solid #555; padding: 6px; text-align: left; word-wrap: break-word;
}
.results-table th:nth-child(3), .results-table td:nth-child(3) { width: 160px; }
.no-results {
  padding: 20px; margin-top: 10px; background-color: #222; color: #f88;
  border: 1px solid #555; border-radius: 4px; font-size: 1.1em; text-align: center;
}
.file-upload .upload-text, .file-upload .file-caption { visibility: hidden !important; }
.file-upload .upload-text::after, .file-upload .file-caption::after {
  content: "Arraste seus PDFs aqui ou clique para enviar";
  visibility: visible !important; display: block; text-align: center; color: #ccc; margin-top: 10px;
}
"""


def strip_accents(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def make_pattern(term: str) -> re.Pattern:
    t = re.escape(strip_accents(term))
    t = t.replace(r"\ ", r"\s+")
    return re.compile(rf"\b{t}\b", flags=re.IGNORECASE)


def _search_one(
    pdf_path: str, names: List[str], chars_before: int, chars_after: int
) -> List[Tuple[str, int, str, str, str]]:
    file_name = os.path.basename(pdf_path)
    results = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                raw = re.sub(r"\s+", " ", page.extract_text() or "")
                sc, mapping = [], []
                for i, ch in enumerate(raw):
                    if ch.isspace():
                        if not sc or sc[-1] != " ":
                            sc.append(" ")
                            mapping.append(i)
                    else:
                        sc.append(strip_accents(ch))
                        mapping.append(i)
                stripped = "".join(sc)

                for name in names:
                    for m in make_pattern(name).finditer(stripped):
                        s, e = m.span()
                        rs, re_ = mapping[s], mapping[e - 1] + 1
                        ss = max(0, rs - chars_before)
                        se = min(len(raw), re_ + chars_after)
                        raw_snip = raw[ss:se]
                        rel_s = rs - ss
                        rel_e = rel_s + (re_ - rs)
                        # escape before injecting highlight span to prevent XSS
                        highlighted = (
                            html_lib.escape(raw_snip[:rel_s])
                            + '<span style="background-color:#ffe066;color:#000">'
                            + html_lib.escape(raw_snip[rel_s:rel_e])
                            + "</span>"
                            + html_lib.escape(raw_snip[rel_e:])
                        )
                        results.append((file_name, page_num, highlighted, name, raw_snip))
    except Exception as exc:
        results.append((
            file_name, 0,
            f'<span style="color:#f88">Erro: {html_lib.escape(str(exc))}</span>',
            "—", "",
        ))
    return results


def search_pdfs(
    pdf_paths: List[str], names: List[str], chars_before: int, chars_after: int
) -> List[Tuple[str, int, str, str, str]]:
    all_results: List[Tuple[str, int, str, str, str]] = []
    with ThreadPoolExecutor(max_workers=min(4, len(pdf_paths))) as ex:
        futures = {
            ex.submit(_search_one, p, names, chars_before, chars_after): p
            for p in pdf_paths
        }
        for f in as_completed(futures):
            all_results.extend(f.result())
    all_results.sort(key=lambda r: (r[0], r[1]))
    return all_results


def _build_excel(matches: List[Tuple[str, int, str, str, str]]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    fill = PatternFill("solid", fgColor="1F4E79")
    bold_white = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(["Arquivo", "Página", "Termo", "Trecho"], 1):
        c = ws.cell(row=1, column=col, value=header)
        c.fill = fill
        c.font = bold_white
    for i, (fname, page, _, term, raw_snip) in enumerate(matches, 2):
        ws.cell(row=i, column=1, value=fname)
        ws.cell(row=i, column=2, value=page)
        ws.cell(row=i, column=3, value=term)
        ws.cell(row=i, column=4, value=raw_snip)
    for col, width in zip("ABCD", [35, 8, 25, 90]):
        ws.column_dimensions[col].width = width
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="busca_")
    wb.save(tmp.name)
    return tmp.name


def gradio_search(pdf_files, names_txt, chars_before, chars_after):
    no_file = gr.update(value=None, visible=False)
    if not pdf_files:
        return "", "<div class='no-results'>Envie ao menos um PDF antes de pesquisar.</div>", no_file
    pdf_paths = [f.name for f in pdf_files]
    names = [n.strip() for n in names_txt.splitlines() if n.strip()]
    if not names:
        return "", "<div class='no-results'>Digite ao menos um termo para buscar.</div>", no_file

    matches = search_pdfs(pdf_paths, names, int(chars_before), int(chars_after))
    valid = [m for m in matches if m[1] != 0]

    if not valid:
        return "", "<div class='no-results'>Nenhum resultado encontrado para os termos informados.</div>", no_file

    termo_counts, termo_pages, arquivo_counts = {}, {}, {}
    for fn, pg, _, t, _ in valid:
        termo_counts[t] = termo_counts.get(t, 0) + 1
        termo_pages.setdefault(t, set()).add(f"{fn} p.{pg}")
        arquivo_counts[fn] = arquivo_counts.get(fn, 0) + 1

    stats_md = (
        f"### Estatísticas de Busca\n"
        f"- **Arquivos pesquisados:** {len(pdf_paths)}\n"
        f"- **Total de ocorrências:** {len(valid)}\n\n"
        "#### Por arquivo\n\n| Arquivo | Ocorrências |\n|---------|:-----------:|\n"
    )
    for fname, cnt in sorted(arquivo_counts.items()):
        stats_md += f"| {html_lib.escape(fname)} | {cnt} |\n"
    stats_md += "\n#### Por termo\n\n| Termo | Total | Ocorrências |\n|-------|:-----:|:-------|\n"
    for term, cnt in termo_counts.items():
        locs = ", ".join(sorted(termo_pages[term]))
        stats_md += f"| {html_lib.escape(term)} | {cnt} | {locs} |\n"

    table_html = (
        "<h3>Resultados Encontrados</h3>"
        "<div id='table-container'>"
        "<table class='results-table'><thead><tr>"
        "<th>Arquivo</th><th>Página</th><th>Termo</th><th>Trecho</th>"
        "</tr></thead><tbody>"
    )
    for fn, pg, snippet, term, _ in matches:
        table_html += (
            f"<tr><td>{html_lib.escape(fn)}</td><td>{pg}</td>"
            f"<td>{html_lib.escape(term)}</td><td>{snippet}</td></tr>"
        )
    table_html += "</tbody></table></div>"

    excel_path = _build_excel(valid)
    return stats_md, table_html, gr.update(value=excel_path, visible=True)


with gr.Blocks() as demo:
    saved_state = gr.BrowserState(
        default_value=["", 200, 200],
        storage_key="pdf_search_v2_settings",
        secret="busca_agil_v2_secret",
    )

    with gr.Row(elem_classes="main-row"):
        with gr.Column(elem_id="controls-col"):
            gr.Markdown("## 🔎 Busca Ágil em PDF v2")
            gr.Markdown(
                "Bem‑vind@ ao **Busca Ágil em PDF v2**, sua ferramenta prática para encontrar "
                "rapidamente nomes ou palavras‑chave em **um ou vários** documentos PDF de uma só vez."
            )
            with gr.Accordion("Como Usar", open=False):
                gr.Markdown(
                    """
1. Envie um ou mais PDFs (pode arrastar vários de uma vez).
2. Digite os termos que deseja buscar (um por linha).
3. (Opcional) Ajuste quantos caracteres de contexto exibir antes e depois de cada ocorrência.
4. Clique em **Pesquisar** e confira as estatísticas e os trechos realçados.
5. Após a busca, baixe os resultados em Excel pelo botão que aparece automaticamente.

Dúvidas ou sugestões? Entre em contato com Paloma Alves da Coordenadoria de Governança de Dados.
                    """
                )
            pdf_input = gr.File(
                label="Enviar PDFs",
                file_count="multiple",
                file_types=[".pdf"],
            )
            names_input = gr.Textbox(
                label="Termos (um por linha)", lines=4, elem_id="names_input"
            )
            with gr.Accordion("Configurações do Contexto", open=False):
                chars_before = gr.Number(
                    label="Caracteres antes do termo pesquisado",
                    value=200, precision=0, elem_id="before_input",
                )
                chars_after = gr.Number(
                    label="Caracteres depois do termo pesquisado",
                    value=200, precision=0, elem_id="after_input",
                )
            search_btn = gr.Button("Pesquisar", variant="primary")
            download_file = gr.File(label="Baixar resultados em Excel", visible=False)
            stats_out = gr.Markdown("", label="Estatísticas de Busca")

        with gr.Column(elem_id="results-col"):
            results_html = gr.HTML("<p>Aguardando pesquisa...</p>")

    demo.load(
        lambda s: (s[0], s[1], s[2]),
        inputs=[saved_state],
        outputs=[names_input, chars_before, chars_after],
    )

    @gr.on(
        [names_input.change, chars_before.change, chars_after.change],
        inputs=[names_input, chars_before, chars_after],
        outputs=[saved_state],
    )
    def save_state(n, b, a):
        return [n, b, a]

    search_btn.click(
        fn=gradio_search,
        inputs=[pdf_input, names_input, chars_before, chars_after],
        outputs=[stats_out, results_html, download_file],
        show_progress=True,
    )

if __name__ == "__main__":
    demo.launch(css=css)
